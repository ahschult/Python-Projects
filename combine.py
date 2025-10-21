import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path

def combine_swim_rankings(file1_path, file2_path, output_path):
    """
    Combine two Excel files containing swim rankings and sort by time (column J).
    Supports both .xls and .xlsx formats.
    
    Args:
        file1_path: Path to first Excel file
        file2_path: Path to second Excel file
        output_path: Path for output Excel file
    """
    
    # Read all sheet names from both files (pandas handles both .xls and .xlsx)
    file1_sheets = pd.ExcelFile(file1_path).sheet_names
    file2_sheets = pd.ExcelFile(file2_path).sheet_names
    
    # Get all unique sheet names
    all_sheets = set(file1_sheets + file2_sheets)
    
    print(f"Processing {len(all_sheets)} unique events...")
    
    # Create output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet
    
    for sheet_name in sorted(all_sheets):
        print(f"  Processing: {sheet_name}")
        
        # Collect data from both files
        dfs = []
        
        if sheet_name in file1_sheets:
            df1 = pd.read_excel(file1_path, sheet_name=sheet_name)
            dfs.append(df1)
        
        if sheet_name in file2_sheets:
            df2 = pd.read_excel(file2_path, sheet_name=sheet_name)
            dfs.append(df2)
        
        # Combine dataframes
        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            
            # Sort by column J (time in seconds) - fastest to slowest (ascending)
            # Handle the case where column J might have different names
            if len(combined_df.columns) >= 10:  # Ensure column J exists (index 9)
                time_col = combined_df.columns[9]  # Column J is index 9
                
                # Convert time column to numeric, coercing errors to NaN
                combined_df[time_col] = pd.to_numeric(combined_df[time_col], errors='coerce')
                
                # Sort by time (NaN values will be placed at the end)
                combined_df = combined_df.sort_values(by=time_col, ascending=True, na_position='last')
            
            # Create new sheet in output workbook
            ws_out = wb_out.create_sheet(title=sheet_name)
            
            # Write data to sheet
            for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_out.cell(row=r_idx, column=c_idx, value=value)
    
    # Save output file
    wb_out.save(output_path)
    print(f"\nCombined file saved to: {output_path}")
    print(f"Total events processed: {len(all_sheets)}")

if __name__ == "__main__":
    # Get the directory where the script is located
    script_dir = Path(__file__).parent
    
    # Find all Excel files in the same directory
    excel_files = list(script_dir.glob("*.xlsx")) + list(script_dir.glob("*.xls"))
    
    # Filter out any files that start with '~' (temporary Excel files)
    excel_files = [f for f in excel_files if not f.name.startswith('~')]
    
    if len(excel_files) != 2:
        print(f"Error: Expected exactly 2 Excel files in the script directory.")
        print(f"Found {len(excel_files)} file(s):")
        for f in excel_files:
            print(f"  - {f.name}")
        print("\nPlease ensure there are exactly 2 Excel files (.xlsx or .xls) in the same directory as this script.")
        exit(1)
    
    file1 = str(excel_files[0])
    file2 = str(excel_files[1])
    
    # Create output filename: combined_[first_filename].xlsx
    file1_stem = excel_files[0].stem  # Get filename without extension
    output = str(script_dir / f"combined_{file1_stem}.xlsx")
    
    print(f"Found 2 Excel files:")
    print(f"  File 1: {excel_files[0].name}")
    print(f"  File 2: {excel_files[1].name}")
    print(f"  Output: combined_{file1_stem}.xlsx\n")
    
    try:
        combine_swim_rankings(file1, file2, output)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        print("Please ensure both Excel files exist in the current directory.")
    except Exception as e:
        print(f"Error processing files: {e}")
        import traceback
        traceback.print_exc()