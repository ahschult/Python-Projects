import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from pathlib import Path
from collections import defaultdict
import re

def combine_swim_rankings(file1_path, file2_path, output_path):
    """
    Combine two Excel files containing swim rankings and sort by time (column J).
    Supports both .xls and .xlsx formats.
    
    Args:
        file1_path: Path to first Excel file
        file2_path: Path to second Excel file
        output_path: Path for output Excel file
    """
    
    # Read all sheet names from both files (pandas handles both .xls and .xlsx)
    file1_sheets = pd.ExcelFile(file1_path).sheet_names
    file2_sheets = pd.ExcelFile(file2_path).sheet_names
    
    # Get all unique sheet names
    all_sheets = set(file1_sheets + file2_sheets)
    
    print(f"    Processing {len(all_sheets)} unique events...")
    
    # Create output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet
    
    for sheet_name in sorted(all_sheets):
        print(f"      - {sheet_name}")
        
        # Collect data from both files
        dfs = []
        
        if sheet_name in file1_sheets:
            df1 = pd.read_excel(file1_path, sheet_name=sheet_name)
            dfs.append(df1)
        
        if sheet_name in file2_sheets:
            df2 = pd.read_excel(file2_path, sheet_name=sheet_name)
            dfs.append(df2)
        
        # Combine dataframes
        if dfs:
            combined_df = pd.concat(dfs, ignore_index=True)
            
            # Sort by column J (time in seconds) - fastest to slowest (ascending)
            # Handle the case where column J might have different names
            if len(combined_df.columns) >= 10:  # Ensure column J exists (index 9)
                time_col = combined_df.columns[9]  # Column J is index 9
                
                # Convert time column to numeric, coercing errors to NaN
                combined_df[time_col] = pd.to_numeric(combined_df[time_col], errors='coerce')
                
                # Sort by time (NaN values will be placed at the end)
                combined_df = combined_df.sort_values(by=time_col, ascending=True, na_position='last')
            
            # Create new sheet in output workbook
            ws_out = wb_out.create_sheet(title=sheet_name)
            
            # Write data to sheet
            for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_out.cell(row=r_idx, column=c_idx, value=value)
    
    # Save output file
    wb_out.save(output_path)

def parse_filename(filename):
    """
    Parse filename to extract matching key components.
    Expected format: CAN-PC_YEAR_FORMAT_GENDER_AGEGROUP
    Returns: (province, year, format, gender, agegroup, full_key)
    """
    # Remove file extension
    name = Path(filename).stem
    
    # Pattern: CAN-PC_YEAR_FORMAT_GENDER_AGEGROUP
    # More flexible pattern that accepts SCM, LCM, or other formats
    pattern = r'CAN-([A-Z]{2})_(\d{4})_([A-Z]+)_(Men|Women)_(.+)'
    match = re.match(pattern, name)
    
    if match:
        province, year, format_type, gender, agegroup = match.groups()
        # Create a key without province for matching
        key = f"{year}_{format_type}_{gender}_{agegroup}"
        return province, year, format_type, gender, agegroup, key
    
    return None

def batch_combine_files():
    """
    Find and combine all matching pairs of Excel files in the data directory.
    """
    # Get script directory and data directory
    script_dir = Path(__file__).parent
    data_dir = script_dir / "data"
    output_dir = script_dir / "combined_data"
    
    if not data_dir.exists():
        print(f"Error: 'data' directory not found at {data_dir}")
        print("Please create a 'data' folder in the same directory as this script.")
        return
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(exist_ok=True)
    print(f"Output directory: {output_dir}\n")
    
    # Find all Excel files in data directory
    excel_files = list(data_dir.glob("*.xlsx")) + list(data_dir.glob("*.xls"))
    excel_files = [f for f in excel_files if not f.name.startswith('~')]
    
    print(f"Found {len(excel_files)} Excel files in data directory\n")
    
    # Group files by their matching key
    file_groups = defaultdict(list)
    
    for file_path in excel_files:
        parsed = parse_filename(file_path.name)
        if parsed:
            province, year, scm, gender, agegroup, key = parsed
            file_groups[key].append((province, file_path))
        else:
            print(f"Warning: Could not parse filename: {file_path.name}")
    
    # Process each matching pair
    processed_count = 0
    skipped_count = 0
    
    for key, files in sorted(file_groups.items()):
        if len(files) != 2:
            print(f"Skipping {key}: Found {len(files)} file(s), expected 2")
            for province, fpath in files:
                print(f"  - {fpath.name}")
            skipped_count += 1
            continue
        
        # Sort to ensure consistent ordering (MB before SK alphabetically)
        files.sort(key=lambda x: x[0])
        
        province1, file1 = files[0]
        province2, file2 = files[1]
        
        # Create output filename
        output_filename = f"CAN-{province1}{province2}_{key}.xlsx"
        output_path = output_dir / output_filename
        
        print(f"Combining pair {processed_count + 1}:")
        print(f"  File 1: {file1.name}")
        print(f"  File 2: {file2.name}")
        print(f"  Output: {output_filename}")
        
        try:
            combine_swim_rankings(str(file1), str(file2), str(output_path))
            print(f"  ✓ Success!\n")
            processed_count += 1
        except Exception as e:
            print(f"  ✗ Error: {e}\n")
            import traceback
            traceback.print_exc()
    
    print(f"\n{'='*60}")
    print(f"Summary:")
    print(f"  Successfully combined: {processed_count} pairs")
    print(f"  Skipped: {skipped_count} files/groups")
    print(f"  Output location: {output_dir}")
    print(f"{'='*60}")

if __name__ == "__main__":
    try:
        batch_combine_files()
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()